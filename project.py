from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time

import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
import random

# firefox 버전
profile = webdriver.FirefoxProfile()
profile.set_preference('general.useragent.override', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0')
profile.set_preference("network.proxy.type", 1)
profile.set_preference("network.proxy.socks", "127.0.0.1")
profile.set_preference("network.proxy.socks_port", 9050)

path = "/Users/taehyung/anaconda3/envs/study/geckodriver"
driver = webdriver.Firefox(firefox_profile=profile, executable_path=path)

# 네이버 지식인 크롤링
# keyword에 크롤링하고 싶은 단어 선택. space 는 + 로 치환
def get_keyword(text):
    return text.replace(" ", "%20")

# 정렬 방식 선택
# 1: 추천순
# 2: 최신순
# 기타: 정확도 순
def sort_kind(index):
    # 추천
    if index == 1:
        return 'vcount'
    # 최신순
    elif index == 2:
        return 'date'
    # 정확도
    else:
        return 'none'


keyword = '자해 자살'
driver.get('https://kin.naver.com/search/list.nhn?query=' + get_keyword(keyword))
driver.implicitly_wait(1000)

page_index = 1
# 크롤링 시작 일자
f = '2019.09.01'
# 크롤링 종료 일자
t = '2019.10.30'
period_txt = "&period=" + f + ".%7C" + t + "."

_sort_kind = sort_kind(2)
date = str(datetime.now()).replace('.', '_')
date = date.replace(' ', '_')

# URL 저장
f = open("result/url_list" + "_" + keyword.replace(' ', '+') + "_" + date + ".txt", 'w')
page_url = []
while True:
    driver.implicitly_wait(random.randint(10, 1000))
    driver.get('https://kin.naver.com/search/list.nhn?' + "&sort=" + _sort_kind + '&query=' + get_keyword(keyword) + period_txt + "&section=kin" + "&page=" + str(page_index))
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    tags = soup.find_all('a', class_="_nclicks:kin.txt _searchListTitleAnchor")
    for tag in tags:
        url = str(tag).split(' ')[3]
        url = url.replace('href=', "")
        url = url.replace('"', "")
        url = url.replace('amp;', '')
        page_url.append(url)
        f.write(url + "\n")

    post_number = driver.find_element_by_class_name('number').text
    post_number = str(post_number).replace("(", "")
    post_number = str(post_number).replace(")", "")
    
    current_number = post_number.split('/')[0].split('-')[1]
    current_number = current_number.replace(',', '')
    total_number = post_number.split('/')[1]
    total_number = total_number.replace(',', '')

    if int(current_number) == int(total_number):
        break
    else:
        page_index += 1

filename = 'result/' + keyword.replace(' ', '.') + "_" + date + "_crawling_result.xlsx"

wb = Workbook()
sheet = wb.active
sheet.append(['제목', '질문', '답변'])

for j in range(1, 4):
    sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

for i in page_url:
    driver.implicitly_wait(random.randint(10, 1000))
    driver.get(i)

    title = driver.find_element_by_class_name('title').text
    print(driver.find_element_by_class_name('c-heading__title-inner'))
    if driver.find_element_by_class_name('c-heading__title-inner') != None:
        question_txt = driver.find_element_by_class_name('c-heading__title-inner').text
    else:
        question_txt = ""
    # 답변 리스트
    answer_list = driver.find_elements_by_class_name("se-main-container")
    
    for n, answer in enumerate(answer_list):
        texts = answer.find_elements_by_tag_name('span')
        t = ""
        for i in texts:
            t += i.text

        if n == 0:
            sheet.append([title, question_txt, t])
        else:
            sheet.append(["", "", t])


wb.save(filename)
driver.quit()
